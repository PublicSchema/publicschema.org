"""Tests for CSV and Excel export generation.

TDD: these tests define expected behavior before implementation.
"""

import csv

import pytest
from openpyxl import load_workbook

from build.build import build_vocabulary
from build.export import generate_concept_csv, generate_definition_xlsx, generate_template_xlsx
from tests.conftest import make_concept, make_property, make_vocabulary

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def sample_vocab_result(
    tmp_schema, write_concept, write_property, write_vocabulary
):
    """Build a vocabulary result with a concept that has vocabulary properties."""
    write_vocabulary("enrollment-status.yaml", make_vocabulary(
        id="enrollment-status",
        values=[
            {
                "code": "active",
                "label": {"en": "Active", "fr": "Actif", "es": "Activo"},
                "definition": {"en": "Currently active.", "fr": "Actif.", "es": "Activo."},
            },
            {
                "code": "suspended",
                "label": {"en": "Suspended", "fr": "Suspendu", "es": "Suspendido"},
                "definition": {"en": "Temporarily paused.", "fr": "Suspendu.", "es": "Suspendido."},
            },
        ],
    ))
    write_property("beneficiary.yaml", make_property(
        id="beneficiary", type="concept:Person",
    ))
    write_property("enrollment_status.yaml", make_property(
        id="enrollment_status", vocabulary="enrollment-status",
    ))
    write_property("enrollment_date.yaml", make_property(
        id="enrollment_date", type="date",
    ))
    write_concept("enrollment.yaml", make_concept(
        id="Enrollment",
        domain="sp",
        properties=["beneficiary", "enrollment_status", "enrollment_date"],
        convergence={"system_count": 6, "total_systems": 6, "notes": "Universal."},
    ))
    return build_vocabulary(tmp_schema)


# ---------------------------------------------------------------------------
# CSV tests
# ---------------------------------------------------------------------------

@pytest.fixture
def inheritance_vocab_result(
    tmp_schema, write_concept, write_property,
):
    """Build a vocabulary result with a Group -> Family/Household hierarchy."""
    write_property("name.yaml", make_property(id="name"))
    write_property("group_type.yaml", make_property(id="group_type"))
    write_property("member_count.yaml", make_property(id="member_count", type="integer"))
    write_property("address.yaml", make_property(id="address"))
    write_concept("group.yaml", make_concept(
        id="Group",
        properties=["name", "group_type", "member_count"],
        subtypes=["Family", "Household"],
    ))
    write_concept("family.yaml", make_concept(
        id="Family",
        properties=[],
        supertypes=["Group"],
    ))
    write_concept("household.yaml", make_concept(
        id="Household",
        properties=["address", "member_count"],
        supertypes=["Group"],
    ))
    return build_vocabulary(tmp_schema)


# ---------------------------------------------------------------------------
# Inheritance tests
# ---------------------------------------------------------------------------

class TestPropertyInheritance:
    """Inherited properties from supertypes appear in all export formats."""

    def test_csv_includes_inherited_properties(self, inheritance_vocab_result, tmp_path):
        """Family (no own properties) gets all properties from Group."""
        generate_concept_csv("Family", inheritance_vocab_result, tmp_path)
        csv_path = tmp_path / "Family.csv"
        reader = csv.DictReader(csv_path.open())
        rows = {r["property"]: r for r in reader}
        assert set(rows.keys()) == {"name", "group_type", "member_count"}

    def test_csv_inherited_properties_come_first(self, inheritance_vocab_result, tmp_path):
        """Inherited properties appear before the concept's own properties."""
        generate_concept_csv("Household", inheritance_vocab_result, tmp_path)
        csv_path = tmp_path / "Household.csv"
        reader = csv.DictReader(csv_path.open())
        props = [r["property"] for r in reader]
        # Group's properties first, then Household's own (minus duplicates)
        assert props == ["name", "group_type", "member_count", "address"]

    def test_csv_deduplicates_inherited_properties(self, inheritance_vocab_result, tmp_path):
        """member_count appears in both Group and Household; should appear only once."""
        generate_concept_csv("Household", inheritance_vocab_result, tmp_path)
        csv_path = tmp_path / "Household.csv"
        reader = csv.DictReader(csv_path.open())
        props = [r["property"] for r in reader]
        assert props.count("member_count") == 1

    def test_definition_xlsx_includes_inherited_properties(
        self, inheritance_vocab_result, tmp_path,
    ):
        """Family definition workbook includes Group's properties."""
        generate_definition_xlsx("Family", inheritance_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "Family-definition.xlsx")
        ws = wb["Properties"]
        props = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert set(props) == {"name", "group_type", "member_count"}

    def test_template_xlsx_includes_inherited_properties(
        self, inheritance_vocab_result, tmp_path,
    ):
        """Family template workbook includes Group's properties as columns."""
        generate_template_xlsx("Family", inheritance_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "Family-template.xlsx")
        ws = wb["Data"]
        ids = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        assert set(ids) == {"name", "group_type", "member_count"}

    def test_parent_concept_unaffected(self, inheritance_vocab_result, tmp_path):
        """Group itself still has exactly its own properties."""
        generate_concept_csv("Group", inheritance_vocab_result, tmp_path)
        csv_path = tmp_path / "Group.csv"
        reader = csv.DictReader(csv_path.open())
        props = [r["property"] for r in reader]
        assert props == ["name", "group_type", "member_count"]


class TestConceptCSV:
    def test_csv_has_header_row(self, sample_vocab_result, tmp_path):
        generate_concept_csv("sp/Enrollment", sample_vocab_result, tmp_path)
        csv_path = tmp_path / "sp" / "Enrollment.csv"
        reader = csv.DictReader(csv_path.open())
        assert set(reader.fieldnames) == {
            "property", "type", "cardinality",
            "definition", "vocabulary",
            "maturity", "sensitivity", "category",
            "age_applicability", "valid_instruments",
        }

    def test_csv_has_one_row_per_property(self, sample_vocab_result, tmp_path):
        generate_concept_csv("sp/Enrollment", sample_vocab_result, tmp_path)
        csv_path = tmp_path / "sp" / "Enrollment.csv"
        reader = csv.DictReader(csv_path.open())
        rows = list(reader)
        assert len(rows) == 3

    def test_csv_property_values(self, sample_vocab_result, tmp_path):
        generate_concept_csv("sp/Enrollment", sample_vocab_result, tmp_path)
        csv_path = tmp_path / "sp" / "Enrollment.csv"
        reader = csv.DictReader(csv_path.open())
        rows = {r["property"]: r for r in reader}

        assert rows["enrollment_status"]["vocabulary"] == "enrollment-status"
        assert rows["enrollment_date"]["type"] == "date"

    def test_csv_universal_concept_no_domain_dir(
        self, tmp_schema, write_concept, write_property, tmp_path
    ):
        """Universal concepts (no domain) write CSV directly in root."""
        write_property("name.yaml", make_property(id="name"))
        write_concept("person.yaml", make_concept(
            id="Person", properties=["name"],
        ))
        result = build_vocabulary(tmp_schema)
        generate_concept_csv("Person", result, tmp_path)
        assert (tmp_path / "Person.csv").exists()


# ---------------------------------------------------------------------------
# Metadata columns (CSV + definition XLSX)
# ---------------------------------------------------------------------------

class TestMetadataExport:
    """Metadata fields from the property schema appear in CSV and XLSX exports."""

    @pytest.fixture
    def metadata_vocab_result(
        self, tmp_schema, write_concept, write_property,
    ):
        write_property("difficulty_walking.yaml", make_property(
            id="difficulty_walking",
            maturity="candidate",
            sensitivity="sensitive",
            category="functioning",
            age_applicability=["adult", "child_5_17"],
            valid_instruments=["wg_ss", "wg_es"],
        ))
        write_property("given_name.yaml", make_property(
            id="given_name",
            maturity="normative",
        ))
        write_concept("person.yaml", make_concept(
            id="Person",
            properties=["difficulty_walking", "given_name"],
        ))
        return build_vocabulary(tmp_schema)

    # --- CSV ---

    def test_csv_metadata_values_populated(self, metadata_vocab_result, tmp_path):
        generate_concept_csv("Person", metadata_vocab_result, tmp_path)
        csv_path = tmp_path / "Person.csv"
        reader = csv.DictReader(csv_path.open())
        rows = {r["property"]: r for r in reader}

        row = rows["difficulty_walking"]
        assert row["maturity"] == "candidate"
        assert row["sensitivity"] == "sensitive"
        assert row["category"] == "functioning"
        assert row["age_applicability"] == "adult, child_5_17"
        assert row["valid_instruments"] == "wg_ss, wg_es"

    def test_csv_metadata_empty_when_absent(self, metadata_vocab_result, tmp_path):
        generate_concept_csv("Person", metadata_vocab_result, tmp_path)
        csv_path = tmp_path / "Person.csv"
        reader = csv.DictReader(csv_path.open())
        rows = {r["property"]: r for r in reader}

        row = rows["given_name"]
        assert row["maturity"] == "normative"
        assert row["sensitivity"] == ""
        assert row["category"] == ""
        assert row["age_applicability"] == ""
        assert row["valid_instruments"] == ""

    # --- Definition XLSX ---

    def test_definition_xlsx_has_metadata_columns(self, metadata_vocab_result, tmp_path):
        generate_definition_xlsx("Person", metadata_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "Person-definition.xlsx")
        ws = wb["Properties"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Maturity" in headers
        assert "Sensitivity" in headers
        assert "Category" in headers
        assert "Age Applicability" in headers
        assert "Valid Instruments" in headers

    def test_definition_xlsx_metadata_values(self, metadata_vocab_result, tmp_path):
        generate_definition_xlsx("Person", metadata_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "Person-definition.xlsx")
        ws = wb["Properties"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

        mat_col = headers.index("Maturity") + 1
        sens_col = headers.index("Sensitivity") + 1
        cat_col = headers.index("Category") + 1
        age_col = headers.index("Age Applicability") + 1
        inst_col = headers.index("Valid Instruments") + 1

        # Row 2 = difficulty_walking (has all metadata)
        assert ws.cell(row=2, column=mat_col).value == "candidate"
        assert ws.cell(row=2, column=sens_col).value == "sensitive"
        assert ws.cell(row=2, column=cat_col).value == "functioning"
        assert ws.cell(row=2, column=age_col).value == "adult, child_5_17"
        assert ws.cell(row=2, column=inst_col).value == "wg_ss, wg_es"

    def test_definition_xlsx_metadata_empty_when_absent(self, metadata_vocab_result, tmp_path):
        generate_definition_xlsx("Person", metadata_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "Person-definition.xlsx")
        ws = wb["Properties"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

        # Row 3 = given_name (minimal metadata)
        mat_col = headers.index("Maturity") + 1
        sens_col = headers.index("Sensitivity") + 1
        assert ws.cell(row=3, column=mat_col).value == "normative"
        assert ws.cell(row=3, column=sens_col).value is None

    # --- Template XLSX (should NOT change) ---

    def test_template_xlsx_excludes_metadata(self, metadata_vocab_result, tmp_path):
        """Metadata columns must not appear in the data-entry template."""
        generate_template_xlsx("Person", metadata_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "Person-template.xlsx")
        ws = wb["Data"]
        ids = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        assert set(ids) == {"difficulty_walking", "given_name"}


# ---------------------------------------------------------------------------
# Definition workbook tests
# ---------------------------------------------------------------------------

class TestDefinitionXLSX:
    def test_has_concept_sheet(self, sample_vocab_result, tmp_path):
        generate_definition_xlsx("sp/Enrollment", sample_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "sp" / "Enrollment-definition.xlsx")
        assert "Concept" in wb.sheetnames

    def test_concept_sheet_has_metadata(self, sample_vocab_result, tmp_path):
        generate_definition_xlsx("sp/Enrollment", sample_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "sp" / "Enrollment-definition.xlsx")
        ws = wb["Concept"]
        # Collect all cell values from column A and B
        all_values = []
        for r in range(1, ws.max_row + 1):
            a = ws.cell(row=r, column=1).value
            b = ws.cell(row=r, column=2).value
            if a:
                all_values.append((a, b))
        # Title row contains concept name
        assert any("Enrollment" in str(a) for a, _ in all_values)
        # Domain is shown as readable label
        data = {a: b for a, b in all_values if b is not None}
        assert any("Social Protection" in str(v) for v in data.values())

    def test_has_properties_sheet(self, sample_vocab_result, tmp_path):
        generate_definition_xlsx("sp/Enrollment", sample_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "sp" / "Enrollment-definition.xlsx")
        assert "Properties" in wb.sheetnames

    def test_properties_sheet_has_header_and_rows(self, sample_vocab_result, tmp_path):
        generate_definition_xlsx("sp/Enrollment", sample_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "sp" / "Enrollment-definition.xlsx")
        ws = wb["Properties"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Property" in headers
        assert "Type" in headers
        assert "Definition (EN)" in headers
        # One header row + 3 property rows
        assert ws.max_row == 4

    def test_has_vocabulary_sheets(self, sample_vocab_result, tmp_path):
        generate_definition_xlsx("sp/Enrollment", sample_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "sp" / "Enrollment-definition.xlsx")
        assert "enrollment-status" in wb.sheetnames

    def test_vocabulary_sheet_has_values(self, sample_vocab_result, tmp_path):
        generate_definition_xlsx("sp/Enrollment", sample_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "sp" / "Enrollment-definition.xlsx")
        ws = wb["enrollment-status"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Code" in headers
        assert "Label (EN)" in headers
        # Header + 2 values
        assert ws.max_row == 3

    def test_no_vocabulary_sheet_for_non_vocab_properties(
        self, tmp_schema, write_concept, write_property, tmp_path
    ):
        """Concepts with no vocabulary properties get no vocab sheets."""
        write_property("name.yaml", make_property(id="name"))
        write_concept("person.yaml", make_concept(
            id="Person", properties=["name"],
        ))
        result = build_vocabulary(tmp_schema)
        generate_definition_xlsx("Person", result, tmp_path)
        wb = load_workbook(tmp_path / "Person-definition.xlsx")
        assert wb.sheetnames == ["Concept", "Properties"]


# ---------------------------------------------------------------------------
# Template workbook tests
# ---------------------------------------------------------------------------

class TestTemplateXLSX:
    def test_has_data_sheet(self, sample_vocab_result, tmp_path):
        generate_template_xlsx("sp/Enrollment", sample_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "sp" / "Enrollment-template.xlsx")
        assert "Data" in wb.sheetnames

    def test_row1_has_human_labels(self, sample_vocab_result, tmp_path):
        generate_template_xlsx("sp/Enrollment", sample_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "sp" / "Enrollment-template.xlsx")
        ws = wb["Data"]
        labels = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        # Labels come from the YAML label.en field (set by make_property fixture)
        assert "Beneficiary" in labels
        assert "Enrollment status" in labels
        assert "Enrollment date" in labels

    def test_row2_has_property_ids(self, sample_vocab_result, tmp_path):
        generate_template_xlsx("sp/Enrollment", sample_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "sp" / "Enrollment-template.xlsx")
        ws = wb["Data"]
        ids = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        assert "beneficiary" in ids
        assert "enrollment_status" in ids
        assert "enrollment_date" in ids

    def test_vocabulary_columns_have_data_validation(self, sample_vocab_result, tmp_path):
        generate_template_xlsx("sp/Enrollment", sample_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "sp" / "Enrollment-template.xlsx")
        ws = wb["Data"]
        # Find the enrollment_status column
        ids = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        status_col = ids.index("enrollment_status") + 1
        # Check that data validation exists on the column
        validations = ws.data_validations.dataValidation
        col_letter = ws.cell(row=3, column=status_col).column_letter
        has_validation = any(
            col_letter in str(dv.sqref) for dv in validations
        )
        assert has_validation, "enrollment_status column should have data validation"

    def test_header_cells_have_comments(self, sample_vocab_result, tmp_path):
        generate_template_xlsx("sp/Enrollment", sample_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "sp" / "Enrollment-template.xlsx")
        ws = wb["Data"]
        # Row 1 (label row) should have comments with definitions
        has_comment = False
        for c in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=c).comment is not None:
                has_comment = True
                break
        assert has_comment, "At least one header cell should have a comment"

    def test_rows_3_onward_are_empty(self, sample_vocab_result, tmp_path):
        generate_template_xlsx("sp/Enrollment", sample_vocab_result, tmp_path)
        wb = load_workbook(tmp_path / "sp" / "Enrollment-template.xlsx")
        ws = wb["Data"]
        for c in range(1, ws.max_column + 1):
            assert ws.cell(row=3, column=c).value is None
