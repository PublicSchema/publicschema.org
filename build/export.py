"""Generates CSV and Excel downloads for each concept.

Outputs are written to a directory structure mirroring concept URL paths,
so they can be served as static files at clean URLs
(e.g., /sp/Enrollment.csv, /Person-definition.xlsx).
"""

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# Colour palette
_BLUE_DARK = "1F3864"
_BLUE_MED = "2E75B6"
_BLUE_LIGHT = "D6E4F0"
_BLUE_HEADER = "D9E1F2"
_GREY_LABEL = "F2F2F2"
_GREY_ALT_ROW = "F7F8FA"
_GREY_TEXT = "808080"
_GREY_BORDER = "B4C6E7"

# Reusable styles
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=_BLUE_DARK)
SUBTITLE_FONT = Font(name="Calibri", size=10, color=_GREY_TEXT)
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color=_BLUE_MED)
LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="333333")
VALUE_FONT = Font(name="Calibri", size=10, color="333333")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=_BLUE_MED, end_color=_BLUE_MED, fill_type="solid")
LABEL_FILL = PatternFill(start_color=_GREY_LABEL, end_color=_GREY_LABEL, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=_GREY_ALT_ROW, end_color=_GREY_ALT_ROW, fill_type="solid")
ID_ROW_FONT = Font(name="Calibri", color=_GREY_TEXT, size=9)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
LABEL_ALIGNMENT = Alignment(vertical="top")
THIN_BORDER_BOTTOM = Border(
    bottom=Side(style="thin", color=_GREY_BORDER),
)

DOMAIN_LABELS = {
    "sp": "Social Protection",
    "edu": "Education",
    "health": "Health",
    "crvs": "Civil Registration and Vital Statistics",
}


def _concept_dir(concept: dict, base_dir: Path) -> Path:
    """Return the output directory for a concept, creating it if needed."""
    domain = concept.get("domain")
    if domain:
        out_dir = base_dir / domain
    else:
        out_dir = base_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def _human_label(prop_id: str) -> str:
    """Convert a property ID to a human-readable label.

    'enrollment_status' -> 'Enrollment status'
    """
    return prop_id.replace("_", " ").capitalize()


def _resolve_properties(
    concept: dict, all_properties: dict, all_concepts: dict | None = None,
) -> list[dict]:
    """Resolve a concept's property references to full property dicts.

    When all_concepts is provided, inherited properties from supertypes
    are included (parent properties first, then the concept's own),
    with duplicates removed.
    """
    seen: set[str] = set()
    resolved: list[dict] = []

    def _add_props(entries):
        for entry in entries:
            prop_id = entry["id"] if isinstance(entry, dict) else entry
            if prop_id not in seen and prop_id in all_properties:
                seen.add(prop_id)
                resolved.append(all_properties[prop_id])

    # Walk supertypes first (inherited properties come before own)
    if all_concepts:
        for parent_id in concept.get("supertypes", []):
            if parent_id in all_concepts:
                parent = all_concepts[parent_id]
                _add_props(parent.get("properties", []))

    _add_props(concept.get("properties", []))
    return resolved


def _clean_text(value: str | None) -> str:
    """Strip trailing whitespace/newlines from YAML multi-line strings."""
    if not value:
        return ""
    return value.strip()


def _style_header_row(ws, row: int, max_col: int):
    """Apply bold font and fill to a table header row."""
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")


def _style_alternating_rows(ws, start_row: int, end_row: int, max_col: int):
    """Apply alternating row shading for readability."""
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).fill = ALT_ROW_FILL


def _auto_column_widths(ws, min_width: int = 8, max_width: int = 50):
    """Set column widths based on content, with reasonable bounds."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # For wrapped cells, use first line length
                text = str(cell.value)
                first_line = text.split("\n")[0] if "\n" in text else text
                max_length = max(max_length, len(first_line))
        ws.column_dimensions[col_letter].width = max(
            min_width, min(max_length + 3, max_width),
        )


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def generate_concept_csv(
    concept_id: str, vocab_result: dict, output_dir: Path
):
    """Generate a flat CSV with one row per property for a concept."""
    concept = vocab_result["concepts"][concept_id]
    properties = _resolve_properties(
        concept, vocab_result["properties"], vocab_result["concepts"],
    )
    out_dir = _concept_dir(concept, output_dir)
    csv_path = out_dir / f"{concept_id}.csv"

    fieldnames = [
        "property", "type", "cardinality",
        "definition", "vocabulary",
    ]
    with csv_path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for prop in properties:
            writer.writerow({
                "property": prop["id"],
                "type": prop.get("type", "string"),
                "cardinality": prop.get("cardinality", "single"),
                "definition": (prop.get("definition") or {}).get("en", ""),
                "vocabulary": prop.get("vocabulary") or "",
            })


# ---------------------------------------------------------------------------
# Definition workbook
# ---------------------------------------------------------------------------

def _write_metadata_row(ws, row: int, label: str, value: str, wrap: bool = False):
    """Write a label-value row on the concept metadata sheet."""
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = LABEL_FONT
    label_cell.fill = LABEL_FILL
    label_cell.alignment = LABEL_ALIGNMENT
    label_cell.border = THIN_BORDER_BOTTOM

    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.font = VALUE_FONT
    value_cell.border = THIN_BORDER_BOTTOM
    if wrap:
        value_cell.alignment = WRAP_ALIGNMENT
    else:
        value_cell.alignment = LABEL_ALIGNMENT


def generate_definition_xlsx(
    concept_id: str, vocab_result: dict, output_dir: Path
):
    """Generate a data dictionary workbook for a concept."""
    concept = vocab_result["concepts"][concept_id]
    properties = _resolve_properties(
        concept, vocab_result["properties"], vocab_result["concepts"],
    )
    vocabularies = vocab_result["vocabularies"]
    out_dir = _concept_dir(concept, output_dir)
    xlsx_path = out_dir / f"{concept_id}-definition.xlsx"

    wb = Workbook()

    # --- Sheet 1: Concept metadata ---
    ws = wb.active
    ws.title = "Concept"
    ws.sheet_properties.tabColor = _BLUE_MED

    # Hide gridlines for a cleaner look
    ws.sheet_view.showGridLines = False

    # Title row: concept name, large and prominent
    domain = concept.get("domain")
    domain_label = DOMAIN_LABELS.get(domain, domain) if domain else None
    title_prefix = f"{domain_label}: " if domain_label else ""
    title_cell = ws.cell(row=1, column=1, value=f"{title_prefix}{concept_id}")
    title_cell.font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 36

    # Subtitle row: branding
    sub_cell = ws.cell(row=2, column=1, value="PublicSchema.org: Data Dictionary")
    sub_cell.font = SUBTITLE_FONT
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 18

    # Blank spacer row
    ws.row_dimensions[3].height = 10

    # --- Identity section ---
    row = 4
    section_cell = ws.cell(row=row, column=1, value="Identity")
    section_cell.font = SECTION_FONT
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    _write_metadata_row(ws, row, "URI", concept.get("uri", ""))
    row += 1
    _write_metadata_row(
        ws, row, "Domain",
        domain_label or "Universal (cross-domain)",
    )
    row += 1
    _write_metadata_row(ws, row, "Maturity", concept.get("maturity", ""))
    row += 1
    _write_metadata_row(
        ws, row, "Properties",
        f"{len(properties)} fields defined",
    )
    row += 1

    # Blank spacer
    row += 1

    # --- Definitions section ---
    definition = concept.get("definition", {})
    section_cell = ws.cell(row=row, column=1, value="Definitions")
    section_cell.font = SECTION_FONT
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    _write_metadata_row(ws, row, "English", _clean_text(definition.get("en")), wrap=True)
    row += 1
    _write_metadata_row(ws, row, "French", _clean_text(definition.get("fr")), wrap=True)
    row += 1
    _write_metadata_row(ws, row, "Spanish", _clean_text(definition.get("es")), wrap=True)
    row += 1

    # Blank spacer
    row += 1

    # --- Convergence section ---
    convergence = concept.get("convergence") or {}
    if convergence:
        section_cell = ws.cell(row=row, column=1, value="Convergence")
        section_cell.font = SECTION_FONT
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        count = convergence.get("system_count", "?")
        total = convergence.get("total_systems", "?")
        _write_metadata_row(ws, row, "Systems", f"{count} of {total} systems mapped")
        row += 1

        notes = _clean_text(convergence.get("notes"))
        if notes:
            _write_metadata_row(ws, row, "Notes", notes, wrap=True)
            row += 1

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 80

    # --- Sheet 2: Properties ---
    ws_props = wb.create_sheet("Properties")
    ws_props.sheet_properties.tabColor = _BLUE_DARK
    prop_headers = [
        "Property", "Type", "Cardinality",
        "Definition (EN)", "Definition (FR)", "Definition (ES)",
        "Vocabulary",
    ]
    for c, header in enumerate(prop_headers, start=1):
        ws_props.cell(row=1, column=c, value=header)
    _style_header_row(ws_props, 1, len(prop_headers))
    ws_props.freeze_panes = "A2"

    for row_idx, prop in enumerate(properties, start=2):
        defn = prop.get("definition") or {}
        ws_props.cell(row=row_idx, column=1, value=prop["id"]).font = VALUE_FONT
        ws_props.cell(row=row_idx, column=2, value=prop.get("type", "string")).font = VALUE_FONT
        ws_props.cell(row=row_idx, column=3, value=prop.get("cardinality", "single")).font = VALUE_FONT
        for col, lang in [(4, "en"), (5, "fr"), (6, "es")]:
            cell = ws_props.cell(row=row_idx, column=col, value=_clean_text(defn.get(lang)))
            cell.font = VALUE_FONT
            cell.alignment = WRAP_ALIGNMENT
        ws_props.cell(row=row_idx, column=7, value=prop.get("vocabulary") or "").font = VALUE_FONT

    _style_alternating_rows(ws_props, 2, len(properties) + 1, len(prop_headers))
    _auto_column_widths(ws_props)

    # --- Sheet 3+: Vocabulary sheets ---
    seen_vocabs = set()
    for prop in properties:
        vocab_id = prop.get("vocabulary")
        if vocab_id and vocab_id in vocabularies and vocab_id not in seen_vocabs:
            seen_vocabs.add(vocab_id)
            vocab = vocabularies[vocab_id]
            # Excel sheet names can't contain '/'; use the bare vocab id.
            ws_vocab = wb.create_sheet(vocab["id"])

            vocab_headers = ["Code", "Label (EN)", "Label (FR)", "Label (ES)", "Definition (EN)"]
            for c, header in enumerate(vocab_headers, start=1):
                ws_vocab.cell(row=1, column=c, value=header)
            _style_header_row(ws_vocab, 1, len(vocab_headers))
            ws_vocab.freeze_panes = "A2"

            for row_idx, value in enumerate(vocab.get("values", []), start=2):
                label = value.get("label") or {}
                defn = value.get("definition") or {}
                ws_vocab.cell(row=row_idx, column=1, value=value["code"]).font = VALUE_FONT
                ws_vocab.cell(row=row_idx, column=2, value=label.get("en", "")).font = VALUE_FONT
                ws_vocab.cell(row=row_idx, column=3, value=label.get("fr", "")).font = VALUE_FONT
                ws_vocab.cell(row=row_idx, column=4, value=label.get("es", "")).font = VALUE_FONT
                cell = ws_vocab.cell(row=row_idx, column=5, value=_clean_text(defn.get("en")))
                cell.font = VALUE_FONT
                cell.alignment = WRAP_ALIGNMENT

            num_values = len(vocab.get("values", []))
            _style_alternating_rows(ws_vocab, 2, num_values + 1, len(vocab_headers))
            _auto_column_widths(ws_vocab)

    wb.save(xlsx_path)


# ---------------------------------------------------------------------------
# Template workbook
# ---------------------------------------------------------------------------

def generate_template_xlsx(
    concept_id: str, vocab_result: dict, output_dir: Path
):
    """Generate a ready-to-fill data entry workbook for a concept."""
    concept = vocab_result["concepts"][concept_id]
    properties = _resolve_properties(
        concept, vocab_result["properties"], vocab_result["concepts"],
    )
    vocabularies = vocab_result["vocabularies"]
    out_dir = _concept_dir(concept, output_dir)
    xlsx_path = out_dir / f"{concept_id}-template.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.sheet_properties.tabColor = _BLUE_MED

    # Hidden reference sheet for vocabulary values (avoids the 255-char
    # limit on inline data validation formulas in Excel).
    ws_ref = None
    ref_col = 0

    for col_idx, prop in enumerate(properties, start=1):
        col_letter = get_column_letter(col_idx)

        # Row 1: human-readable label
        label_cell = ws.cell(row=1, column=col_idx, value=_human_label(prop["id"]))
        label_cell.font = HEADER_FONT
        label_cell.fill = HEADER_FILL

        # Add comment with definition
        defn = (prop.get("definition") or {}).get("en", "")
        if defn:
            label_cell.comment = Comment(defn, "PublicSchema")

        # Row 2: property ID
        id_cell = ws.cell(row=2, column=col_idx, value=prop["id"])
        id_cell.font = ID_ROW_FONT

        # Data validation for vocabulary properties
        vocab_id = prop.get("vocabulary")
        if vocab_id and vocab_id in vocabularies:
            vocab = vocabularies[vocab_id]
            codes = [v["code"] for v in vocab.get("values", [])]
            if codes:
                # Write codes to a hidden reference sheet and point
                # the validation formula at that range.
                if ws_ref is None:
                    ws_ref = wb.create_sheet("_values")
                    ws_ref.sheet_state = "hidden"
                ref_col += 1
                ref_letter = get_column_letter(ref_col)
                ws_ref.cell(row=1, column=ref_col, value=prop["id"])
                for r, code in enumerate(codes, start=2):
                    ws_ref.cell(row=r, column=ref_col, value=code)
                last_row = len(codes) + 1
                formula = f"_values!${ref_letter}$2:${ref_letter}${last_row}"

                dv = DataValidation(
                    type="list",
                    formula1=formula,
                    allow_blank=True,
                )
                dv.error = f"Value must be from the {vocab_id} vocabulary."
                dv.errorTitle = f"Invalid {_human_label(prop['id'])}"
                dv.sqref = f"{col_letter}3:{col_letter}1000"
                ws.add_data_validation(dv)

    # Freeze the first two rows (labels + IDs)
    ws.freeze_panes = "A3"
    _auto_column_widths(ws)

    wb.save(xlsx_path)


# ---------------------------------------------------------------------------
# Batch generation
# ---------------------------------------------------------------------------

def generate_all_downloads(vocab_result: dict, output_dir: Path):
    """Generate CSV and Excel downloads for all concepts."""
    output_dir.mkdir(parents=True, exist_ok=True)
    for concept_id in vocab_result["concepts"]:
        generate_concept_csv(concept_id, vocab_result, output_dir)
        generate_definition_xlsx(concept_id, vocab_result, output_dir)
        generate_template_xlsx(concept_id, vocab_result, output_dir)
